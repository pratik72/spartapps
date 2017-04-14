import os, openpyxl
from pymongo import MongoClient
from bson import ObjectId
from datetime import datetime as dt
from openpyxl.styles import Font

working_directory = os.getcwd()
save_path = working_directory + "\\Reports\\PO Orders"
if not os.path.exists(save_path):
    os.makedirs(save_path)
client = MongoClient()
db = client.spartadb
po_collection = db.purchase_orders
user_collection = db.users
workbook = openpyxl.Workbook()
active_sheet = workbook.get_sheet_by_name("Sheet")
date_stamp = str(dt.now().date().day) + '-' + str(dt.now().date().month) + '-' + str(dt.now().date().year)
time_stamp = str(dt.now().hour) + '-' + str(dt.now().minute) + '-' + str(dt.now().second)
save_name = 'PO_Report_' + date_stamp + '_' + time_stamp + '.xlsx'
header_names = ['PO Number', 'Location', 'Supplier', 'CST', 'GST', 'HSN Code', 'VAT', 'Amount', 'Excise',
                'Payment Terms', 'PO Category',	'Product Description', 'Product Name', 'Quantity', 'Rate',
                'Service Tax', 'Username', 'Approved By', 'Division']
font = Font(bold=True)

os.chdir(save_path)
for head in header_names:
    cell = active_sheet.cell(row=1, column=header_names.index(head)+1)
    cell.value = head
    cell.font = font

line = 1
for doc in po_collection.find({}):
    line += 1
    active_sheet.cell(row=line, column=1).value = doc['PO_number']
    active_sheet.cell(row=line, column=2).value = doc['budgets_and_approvals']['location']
    active_sheet.cell(row=line, column=3).value = doc['supplier_name']
    active_sheet.cell(row=line, column=4).value = doc['product_information']['CST']
    active_sheet.cell(row=line, column=5).value = doc['product_information']['GST']
    active_sheet.cell(row=line, column=6).value = doc['product_information']['HSN_code']
    active_sheet.cell(row=line, column=7).value = doc['product_information']['VAT']
    active_sheet.cell(row=line, column=8).value = doc['product_information']['amount']
    active_sheet.cell(row=line, column=9).value = doc['product_information']['excise']
    active_sheet.cell(row=line, column=10).value = doc['product_information']['payment_terms']
    active_sheet.cell(row=line, column=11).value = doc['product_information']['poCategory']
    active_sheet.cell(row=line, column=12).value = doc['product_information']['product_discreption']
    active_sheet.cell(row=line, column=13).value = doc['product_information']['product_name']
    active_sheet.cell(row=line, column=14).value = doc['product_information']['quantity']
    active_sheet.cell(row=line, column=15).value = doc['product_information']['rate']
    active_sheet.cell(row=line, column=16).value = doc['product_information']['service_tax']
    first_name = user_collection.find_one({'_id': ObjectId(doc['vendor_selection']['selected_by'])})['firstName']
    last_name =  user_collection.find_one({'_id': ObjectId(doc['vendor_selection']['selected_by'])})['lastName']
    active_sheet.cell(row=line, column=17).value = doc['userName']
    active_sheet.cell(row=line, column=18).value = first_name + ' ' + last_name
    active_sheet.cell(row=line, column=19).value = doc['vendor_selection']['division']

workbook.save(save_name)
